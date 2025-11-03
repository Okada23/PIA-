import os
import sys
import csv
import json
import sqlite3
import datetime
from sqlite3 import Error
from tabulate import tabulate
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side


if not os.path.exists("ReservasCoworking.db"):
    print("\nNo se encontró base de datos anterior. Se inicia con estado vacío.")
else:
    print("\nSe ha recuperado el estado anterior.")

try:
    with sqlite3.connect("ReservasCoworking.db") as conn:
        cursor = conn.cursor()
        
        cursor.execute("""CREATE TABLE IF NOT EXISTS clientes (
            clave INTEGER PRIMARY KEY,
            nombre TEXT NOT NULL,
            apellido TEXT NOT NULL
        )""")
        
        cursor.execute("""CREATE TABLE IF NOT EXISTS salas (
            clave INTEGER PRIMARY KEY,
            nombre TEXT NOT NULL,
            cupo INTEGER NOT NULL
        )""")
        
        cursor.execute("""CREATE TABLE IF NOT EXISTS turno (
            clave_horario INTEGER PRIMARY KEY,
            tipo_turno TEXT NOT NULL
        )""")
        
        cursor.execute("SELECT COUNT(*) FROM turno")
        if cursor.fetchone()[0] == 0:
            cursor.executemany(
                "INSERT INTO turno (tipo_turno) VALUES (?)",
                [("Matutino",), ("Vespertino",), ("Nocturno",)]
            )
        
        cursor.execute("""CREATE TABLE IF NOT EXISTS reserva (
            folio INTEGER PRIMARY KEY,
            fecha TIMESTAMP NOT NULL,
            clave_sala INTEGER NOT NULL,
            turno TEXT NOT NULL,
            clave_cliente INTEGER NOT NULL,
            evento TEXT NOT NULL,
            creado TEXT NOT NULL,
            estado TEXT NOT NULL DEFAULT 'ACTIVA',
            FOREIGN KEY(clave_sala) REFERENCES salas(clave),
            FOREIGN KEY(clave_cliente) REFERENCES clientes(clave)
        )""")
        
        conn.commit()

except Error as e:
    print(e)
except Exception:
    print(f"\nSe produjo el siguiente error: {sys.exc_info()[0]}")


def mostrar_clientes_ordenados():
    """Muestra los clientes de la base de datos ordenados alfabéticamente y permite seleccionar uno."""
    try:
        with sqlite3.connect("ReservasCoworking.db") as conn:
            mi_cursor = conn.cursor()

            while True:
                mi_cursor.execute("SELECT clave, apellido, nombre FROM clientes ORDER BY apellido, nombre")
                clientes = mi_cursor.fetchall()
                
                if not clientes:
                    print("\nNo hay clientes registrados aún.")
                    return None

                print("\n" + "=" * 30)
                print(f"{'LISTA DE CLIENTES':^28}")
                print("=" * 30)
                for clave, apellido, nombre in clientes:
                    print(f"{clave} - {apellido}, {nombre}")

                respuesta_clave_cliente = input("\nIngrese la clave del cliente (o escriba 'EXIT' para cancelar): ").strip().upper()

                if respuesta_clave_cliente == "EXIT":
                    print("\nOperación cancelada.")
                    return None

                for clave, apellido, nombre in clientes:
                    if str(clave) == respuesta_clave_cliente:
                        print(f"\nCliente seleccionado: {apellido.upper()}, {nombre.upper()}")
                        return clave, apellido, nombre
          
                print("\nNo existe esa clave. Se muestra la lista nuevamente:")

    except Error as e:
        print(e)
    except Exception:
        print(f"\nSe produjo el siguiente error: {sys.exc_info()[0]}")


def seleccionar_fecha_reservacion():
    while True:
        fecha_reserva = input("\nIngrese la fecha de la reserva (mm-dd-aaaa) o EXIT para cancelar: ").strip()

        if fecha_reserva == "":
            print("\nNo se puede dejar vacio el dato.")
            continue
        elif fecha_reserva.upper() == "EXIT":
            print ("Operación cancelada.")
            break

        try:
            fecha_convertida = datetime.datetime.strptime(fecha_reserva, "%m-%d-%Y").date()
            fecha_minima = datetime.date.today() + datetime.timedelta(days=2)
            
            if fecha_convertida < fecha_minima:
                print(f"\nLa fecha debe ser al menos dos días posteriores a hoy ({datetime.date.today().strftime('%m-%d-%Y')}).")
                continue
            
            if fecha_convertida.weekday() == 6: 
                fecha_propuesta = fecha_convertida + datetime.timedelta(days=1)
                print("\nNo se pueden realizar reservaciones en domingo.")
                
                while True:
                    respuesta = input(f"\n¿Desea reservar el lunes siguiente ({fecha_propuesta.strftime('%m-%d-%Y')})? (S/N): ").strip().upper()
                    
                    if respuesta == 'S':
                        if fecha_propuesta < fecha_minima:
                            print(f"\nEl lunes propuesto ({fecha_propuesta.strftime('%m-%d-%Y')}) tampoco cumple los 2 días de anticipación.")
                            break
                        return fecha_propuesta
                    
                    elif respuesta == 'N':
                        print("\nIngrese otra fecha que no sea en domingo.")
                        break 
                    else:
                        print("\nRespuesta no válida. Intente con 'S' o 'N'.")
                continue

            return fecha_convertida

        except ValueError:
            print("\nFormato de fecha incorrecto. Use mm-dd-aaaa.")


def seleccionar_turno():
    """Permite seleccionar un turno válido desde la tabla turno."""
    try:
        with sqlite3.connect ("ReservasCoworking.db") as conn:
            cursor = conn.cursor()

            while True:
                cursor.execute("SELECT clave_horario, tipo_turno FROM turno")
                filas = cursor.fetchall()
                
                if not filas:
                    print("\nNo hay turnos definidos en la base de datos.")
                    return None

                TURNOS = {str(clave): descripcion for clave, descripcion in filas}

                print("\n" + "=" * 40)
                print(f"{'TURNOS DISPONIBLES':^38}")
                print("=" * 40)
                for clave, descripcion in TURNOS.items():
                    print(f"{clave} - {descripcion}")

                respuesta_turno = input("\nIngrese la clave del turno (o escriba 'EXIT' para volver): ").strip().upper()

                if respuesta_turno == "EXIT":
                    print("Regresando al menú...")
                    return None
                

                if respuesta_turno not in TURNOS:
                    print("\nOpción de turno inválida. Intente de nuevo.")
                    continue

                return TURNOS[respuesta_turno]

    except Error as e:
        print(e)
    except Exception:
        print(f"\nSe produjo el siguiente error: {sys.exc_info()[0]}")


def seleccionar_sala(fecha_reserva, turno_seleccionado):
    """Permite seleccionar una sala disponible para la fecha y turno indicados."""
    try:
        fecha_texto_sql = fecha_reserva.strftime("%Y-%m-%d")
        fecha_texto_usuario = fecha_reserva.strftime("%m-%d-%Y")

        with sqlite3.connect("ReservasCoworking.db") as conn:
            mi_cursor = conn.cursor()
            mi_cursor.execute(
                "SELECT clave, nombre, cupo FROM salas WHERE clave NOT IN ("
                "SELECT clave_sala FROM reserva WHERE fecha = ? AND turno = ?)",
                (fecha_texto_sql, turno_seleccionado)
            )
            salas_disponibles = mi_cursor.fetchall()

            if not salas_disponibles:
                print(f"\nNo hay salas disponibles el {fecha_texto_usuario} en ese turno.")
                return None

            print("\n" + "=" * 50)
            print(f"SALAS DISPONIBLES el {fecha_texto_usuario} en turno {turno_seleccionado}")
            print("=" * 50)
            for clave_sala, nombre_sala, cupo_sala in salas_disponibles:
                print(f"\n{clave_sala} - Sala {nombre_sala} para {cupo_sala} personas")

            while True:
                respuesta_sala = input("\nIngrese la clave de la sala (o escriba 'BACK' para volver a elegir turno): ").strip().upper()
                if respuesta_sala == "BACK":
                    return "BACK"

                for clave_sala, nombre_sala, cupo_sala in salas_disponibles:
                    if str(clave_sala) == respuesta_sala:
                        return clave_sala

                print("\nClave de sala inválida. Intente de nuevo.")

    except Error as e:
        print(e)
    except Exception:
        print(f"\nSe produjo el error {sys.exc_info()[0]}")


def asignar_nombre_evento():
    while True:
        nombre_evento = input("\nIngrese el nombre del evento (o escriba 'BACK' para volver a elegir sala): ").strip()

        if nombre_evento.upper() == "BACK":
            return "BACK"  

        while "  " in nombre_evento:
            nombre_evento = nombre_evento.replace("  ", " ")

        if nombre_evento == "":
            print("\nEl nombre del evento es obligatorio y no puede dejarse vacío.")
        else:
            return nombre_evento
        

def registrar_reserva_de_sala():
    """Unifica funciones para ser colocado en el menú."""
    resultado_cliente = mostrar_clientes_ordenados()
    if not resultado_cliente:
        return
    clave_cliente = resultado_cliente[0]

    fecha_reserva = seleccionar_fecha_reservacion()
    if not fecha_reserva:
        return

    while True:
        turno_seleccionado = seleccionar_turno()
        if not turno_seleccionado:
            return

        while True:
            clave_sala = seleccionar_sala(fecha_reserva, turno_seleccionado)
            if clave_sala == "BACK":
                break  
            elif not clave_sala:
                print("\nNo hay salas disponibles para este turno.")
                break 

            while True:
                nombre_evento_final = asignar_nombre_evento()
                if nombre_evento_final == "BACK":
                    break 

                fecha_texto = fecha_reserva.strftime("%Y-%m-%d")
                fecha_creacion = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                try:
                    with sqlite3.connect("ReservasCoworking.db") as conn:
                        mi_cursor = conn.cursor()

                        mi_cursor.execute("""
                            SELECT COUNT(*) FROM reserva 
                            WHERE fecha = ? AND clave_sala = ? AND turno = ?
                        """, (fecha_texto, clave_sala, turno_seleccionado))
                        existe_reserva = mi_cursor.fetchall()[0][0]

                        if existe_reserva > 0:
                            print("\nLa sala ya está reservada para ese turno y fecha. Intente con otra sala o turno.")
                            break  
                        reserva_nueva = (
                            fecha_texto, clave_sala, turno_seleccionado, clave_cliente,
                            nombre_evento_final.capitalize(), fecha_creacion
                        )
                        mi_cursor.execute("""
                            INSERT INTO reserva (fecha, clave_sala, turno, clave_cliente, evento, creado)
                            VALUES (?, ?, ?, ?, ?, ?)
                        """, reserva_nueva)
                        conn.commit()

                        print(f"\nReserva confirmada con folio: {mi_cursor.lastrowid}")
                    return  

                except Error as e:
                    print(e)
                except Exception:
                    print(f"\nSe produjo el siguiente error: {sys.exc_info()[0]}")


def editar_nombre_de_evento():
    """Permite modificar el nombre de un evento existente dentro de un rango de fechas."""

    while True:
        with sqlite3.connect("ReservasCoworking.db") as conn:
            mi_cursor = conn.cursor()
            mi_cursor.execute("SELECT COUNT(*) FROM reserva")
            total = mi_cursor.fetchone()[0]
            if total == 0:
                print("\nNo hay reservas previamente registradas.")
                return

        while True:
            fecha_inicial = input("\nIngrese la primera fecha (mm-dd-aaaa)o escriba 'EXIT' para cancelar: ").strip()
            if fecha_inicial.upper() == "EXIT":
                    print("\nOperación cancelada.\n")
                    return
            if fecha_inicial == "":
                print("\nNo se puede dejar vacío el dato.")
                continue

            try:
                fecha_inicio = datetime.datetime.strptime(fecha_inicial, "%m-%d-%Y").date()
                break
            except ValueError:
                print("\nFormato incorrecto. Intente de nuevo.")

        while True:
            fecha_final = input("\nIngrese la segunda fecha (mm-dd-aaaa) o escriba 'BACK' para regresar a fecha inicial: ").strip()
            if fecha_final.upper() == "BACK":
                print("\nVolviendo al ingreso de la fecha inicial...\n")
                break

            if fecha_final == "":
                print("\nNo se puede dejar vacío el dato.")
                continue

            try:
                fecha_fin = datetime.datetime.strptime(fecha_final, "%m-%d-%Y").date()
                if fecha_fin <= fecha_inicio:
                    print("\nLa segunda fecha debe ser posterior a la primera.")
                    continue
                break
            except ValueError:
                print("\nFormato incorrecto. Intente de nuevo.")

        if fecha_final.upper() == "BACK":
            continue
        break

    try:
        with sqlite3.connect("ReservasCoworking.db") as conn:
            mi_cursor = conn.cursor()

            fecha_inicio_texto = fecha_inicio.strftime("%Y-%m-%d")
            fecha_fin_texto = fecha_fin.strftime("%Y-%m-%d")

            mi_cursor.execute(
                "SELECT folio, fecha, evento FROM reserva "
                "WHERE fecha BETWEEN ? AND ? AND estado = 'ACTIVA' ORDER BY fecha",
                (fecha_inicio_texto, fecha_fin_texto)
            )
            reservas = mi_cursor.fetchall()

            if not reservas:
                print("\nNo hay reservas en este rango.\n")
                return

            lista_mostrar = [
                [folio, evento, datetime.datetime.strptime(fecha, "%Y-%m-%d").strftime("%m-%d-%Y")]
                for folio, fecha, evento in reservas
            ]
            print(tabulate(lista_mostrar, headers=["Folio", "Evento", "Fecha"], tablefmt="grid"))

            while True:
                folio_a_modificar = input("\nIngrese el folio del evento a modificar (o escriba 'EXIT' para cancelar): ").strip()
                if folio_a_modificar.upper() == "EXIT":
                    print("\nOperación cancelada.\n")
                    return

                try:
                    folio_validacion = int(folio_a_modificar)
                except ValueError:
                    print("\nFolio inválido. Intente de nuevo.")
                    continue

                mi_cursor.execute("SELECT folio FROM reserva WHERE folio = ?", (folio_validacion,))
                if not mi_cursor.fetchall():
                    print("\nFolio no encontrado en el rango. Intente de nuevo.")
                    continue

                nuevo_nombre_evento = input("\nIngrese el nuevo nombre del evento: ").strip()
                if not nuevo_nombre_evento:
                    print("\nEl nombre no puede estar vacío.")
                    continue

                mi_cursor.execute(
                    "UPDATE reserva SET evento = ? WHERE folio = ?",
                    (nuevo_nombre_evento.capitalize(), folio_validacion)
                )
                conn.commit()
                print(f"\nEvento actualizado correctamente: {nuevo_nombre_evento.capitalize()}\n")
                return

    except Error as e:
        print(e)
    except Exception:
        print(f"\nSe produjo el siguiente error: {sys.exc_info()[0]}")


def consultar_reservas_por_fecha():
    """Consulta las reservas en la base de datos por fecha, mostrando sala, cliente, evento y turno."""
    with sqlite3.connect("ReservasCoworking.db") as conn:
            mi_cursor = conn.cursor()
            mi_cursor.execute("SELECT COUNT(*) FROM reserva")
            total = mi_cursor.fetchone()[0]
            if total == 0:
                print("\nNo hay reservas previamente registrados.")
                return
    fecha_consulta = input("\nIngrese la fecha a consultar (mm-dd-aaaa) o presione ENTER para tomar la fecha actual. Si desea cancelar escriba 'EXIT' : ").strip()
    if fecha_consulta.upper() == "EXIT":
                    print("\nOperación cancelada.\n")
                    return

    if not fecha_consulta:
        fecha_consulta_convertida = datetime.date.today()
        print(f"\nAsumiendo fecha actual del sistema: {fecha_consulta_convertida.strftime('%m-%d-%Y')}")
    else:
        try:
            fecha_consulta_convertida = datetime.datetime.strptime(fecha_consulta, "%m-%d-%Y").date()
        except ValueError:
            print("\nFormato incorrecto. Debería ser mm-dd-aaaa.")
            return

    fecha_texto = fecha_consulta_convertida.strftime("%Y-%m-%d")

    try:
        with sqlite3.connect("ReservasCoworking.db") as conn:
            mi_cursor = conn.cursor()
            mi_cursor.execute(
                "SELECT clave_sala, clave_cliente, evento, turno FROM reserva WHERE fecha = ? AND estado = 'ACTIVA'",
                (fecha_texto,)
            )
            reservas = mi_cursor.fetchall()

            if not reservas:
                print("\nNo hay reservas para esa fecha.\n")
                return

            filas_reservas = []

            for clave_sala, clave_cliente, evento, turno in reservas:
                mi_cursor.execute("SELECT nombre FROM salas WHERE clave = ?", (clave_sala,))
                resultado_sala = mi_cursor.fetchone()
                nombre_sala = resultado_sala[0] if resultado_sala else ""

                mi_cursor.execute("SELECT nombre, apellido FROM clientes WHERE clave = ?", (clave_cliente,))
                resultado_cliente = mi_cursor.fetchone()
                if resultado_cliente:
                    nombre_cliente, apellido_cliente = resultado_cliente
                else:
                    nombre_cliente, apellido_cliente = ("", "")

                nombre_cliente_completo = f"{nombre_cliente} {apellido_cliente}".strip()
                filas_reservas.append([nombre_sala, nombre_cliente_completo, evento, turno])

            if filas_reservas:
                headers = ["SALA", "CLIENTE", "EVENTO", "TURNO"]
                tabla = tabulate(filas_reservas, headers=headers, tablefmt="grid")

                ancho_tabla = len(tabla.split("\n")[0])
                print("\n" + "=" * ancho_tabla)
                titulo = f"REPORTE DE RESERVACIONES PARA EL DÍA {fecha_texto}"
                print(titulo.center(ancho_tabla))
                print("=" * ancho_tabla)

                print(tabla)

            exportar_reporte(filas_reservas)

    except Error as e:
        print(e)
    except Exception:
        print(f"\nSe produjo el error: {sys.exc_info()[0]}")


def exportar_reporte(filas):
    if not filas:
        print("\nNo hay datos para exportar.\n")
        return
    
    print("\n" + "=" * 30)
    print(f"{'OPCIONES DE EXPORTACION':^28}")
    print("=" * 30)
    print("1. CSV")
    print("2. JSON")
    print("3. Excel")

    try:
        opcion_exportacion = int(input("\nSeleccione el formato: ").strip())

        if opcion_exportacion == 1:
            with open("reporte.csv", "w", newline="", encoding="utf-8") as archivo_csv:
                escritor_csv = csv.writer(archivo_csv)
                escritor_csv.writerow(["Sala", "Cliente", "Evento", "Turno"])
                for fila_reserva in filas:
                    fila_completa = [valor if valor is not None else "" for valor in fila_reserva]
                    escritor_csv.writerow(fila_completa)

            print("\nReporte exportado a 'reporte.csv'")

        elif opcion_exportacion == 2:
            reporte_json = []
            for fila_reserva in filas:
                fila_completa = [valor if valor is not None else "" for valor in fila_reserva]
                registro = {
                    "Sala": fila_completa[0],
                    "Cliente": fila_completa[1],
                    "Evento": fila_completa[2],
                    "Turno": fila_completa[3]
                }
                reporte_json.append(registro)

            with open("reporte.json", "w", encoding="utf-8") as archivo_json:
                json.dump(reporte_json, archivo_json, ensure_ascii=False, indent=2)

            print("\nReporte exportado a 'reporte.json'")

        elif opcion_exportacion == 3:
            libro_excel = Workbook()
            hoja_excel = libro_excel.active
            hoja_excel.title = "Reservas"
            encabezados = ["Sala", "Cliente", "Evento", "Turno"]

            for col_index, encabezado in enumerate(encabezados, start=1):
                celda = hoja_excel.cell(row=1, column=col_index, value=encabezado)
                celda.font = Font(bold=True)
                celda.alignment = Alignment(horizontal="center")
                celda.border = Border(bottom=Side(style="thick"))

            for row_index, fila_reserva in enumerate(filas, start=2):
                fila_completa = [valor if valor is not None else "" for valor in fila_reserva]
                for col_index, valor_celda in enumerate(fila_completa, start=1):
                    celda = hoja_excel.cell(row=row_index, column=col_index, value=valor_celda)
                    celda.alignment = Alignment(horizontal="center")

            libro_excel.save("reporte.xlsx")
            print("\nReporte exportado a 'reporte.xlsx'")

        else:
            print("Opción inválida.")
        
    except ValueError:
        print("\nSolo se aceptan números enteros de los que están disponibles (1-3).")


def registrar_cliente():
    '''Registra un cliente en la Base de Datos y devuelve una clave única generada automáticamente.'''
    while True: 
        
        while True:
            nombre_cliente = input("\nIngrese el nombre del cliente (Escriba 'EXIT' para salir): ").strip()
            
            if nombre_cliente.upper() == "EXIT":
                print("\nOperación cancelada.")
                return 
            
            if nombre_cliente == "":
                print("\nNo se puede omitir el dato.")
                continue

            if not nombre_cliente.replace(" ", "").isalpha():
                print("\nNo se aceptan números ni caracteres especiales.")
                continue
            break

        while True:
            apellido_cliente = input("\nIngrese el apellido del cliente (o escriba BACK para corregir nombre): ").strip()
            
            if apellido_cliente.upper() == "BACK":
                print("\nRegresando a ingresar nombre...")
                break
            
            if apellido_cliente == "":
                print("\nNo se puede omitir el dato.")
                continue

            if not apellido_cliente.replace(" ", "").isalpha():
                print("\nNo se aceptan números ni caracteres especiales.")
                continue
            break
        
        if apellido_cliente.upper() == "BACK":
            continue
        break 

    cliente_nuevo = (nombre_cliente.upper(), apellido_cliente.upper())

    try:
        with sqlite3.connect("ReservasCoworking.db") as conn:
            mi_cursor = conn.cursor()
            mi_cursor.execute("INSERT INTO clientes (nombre, apellido) VALUES (?,?)", cliente_nuevo)
            print(f"\nCliente registrado correctamente. Clave asignada: {mi_cursor.lastrowid}")

    except Error as e:
        print(e)
    except Exception:
        print(f"\nSe produjo el siguiente error: {sys.exc_info()[0]}")


def registrar_sala():
    while True: 
        
        while True:
            nombre_sala = input("\nIngrese el nombre de la nueva sala (o escriba 'EXIT' para cancelar): ").strip().upper()
            if nombre_sala == "EXIT":
                print("\nOperación cancelada.")
                return 
            if nombre_sala == "":
                print("\nNo se puede omitir el dato.")
                continue
            break 
        
        while True:
            cupo_input = input("\nIngrese el cupo de la sala (o escriba BACK para corregir nombre de sala): ").strip().upper()
            
            if cupo_input == "BACK":
                print("\nRegresando a ingresar nombre de sala...")
                break
                
            try:
                cupo_sala = int(cupo_input) 
                
                if cupo_sala <= 0:
                    print("\nEl cupo debe ser un número entero POSITIVO.")
                    continue
                
                break
                
            except ValueError:
                print("\nEl cupo no es un número entero válido. Intente de nuevo.")
                continue

        if cupo_input == "BACK":
            continue
        break 

    sala_nueva = (nombre_sala, cupo_sala)

    try:
        with sqlite3.connect("ReservasCoworking.db") as conn:
            mi_cursor = conn.cursor()
            mi_cursor.execute("INSERT INTO salas (nombre, cupo) VALUES (?,?)", sala_nueva)
            print(f"\nLa clave asignada fue {mi_cursor.lastrowid}")

    except Error as e:
        print(e)
    except Exception:
        print(f"\nSe produjo el siguiente error: {sys.exc_info()[0]}")


def cancelar_reservas():
    while True:
        with sqlite3.connect("ReservasCoworking.db") as conn:
            mi_cursor = conn.cursor()
            mi_cursor.execute("SELECT COUNT(*) FROM reserva")
            total = mi_cursor.fetchone()[0]
            if total == 0:
                print("\nNo hay reservas previamente registradas.")
                return

        while True:
            fecha_inicial = input("\nIngrese la primera fecha (mm-dd-aaaa) o escriba 'EXIT' para cancelar: ").strip()
            if fecha_inicial.upper() == "EXIT":
                    print("\nOperación cancelada.\n")
                    return
            if fecha_inicial == "":
                print("\nNo se puede dejar vacío el dato.")
                continue
            try:
                fecha_inicio = datetime.datetime.strptime(fecha_inicial, "%m-%d-%Y").date()
                break
            except ValueError:
                print("\nFormato incorrecto. Intente de nuevo.")

        while True:
            fecha_final = input("\nIngrese la segunda fecha (mm-dd-aaaa) o escriba 'BACK' para regresar: ").strip()
            if fecha_final.upper() == "BACK":
                print("\nRegresando al ingreso de la primera fecha...\n")
                break

            if fecha_final == "":
                print("\nNo se puede dejar vacío el dato.")
                continue

            try:
                fecha_fin = datetime.datetime.strptime(fecha_final, "%m-%d-%Y").date()
                if fecha_fin <= fecha_inicio:
                    print("\nLa segunda fecha debe ser posterior a la primera.")
                    continue
                break
            except ValueError:
                print("\nFormato incorrecto. Intente de nuevo.")

        if fecha_final.upper() == "BACK":
            continue
        break

    try:
        with sqlite3.connect("ReservasCoworking.db") as conn:
            mi_cursor = conn.cursor()
            fecha_inicio_texto = fecha_inicio.strftime("%Y-%m-%d")
            fecha_fin_texto = fecha_fin.strftime("%Y-%m-%d")
            mi_cursor.execute(
                "SELECT folio, fecha, evento FROM reserva WHERE fecha BETWEEN ? AND ? AND estado = 'ACTIVA' ORDER BY fecha",
                (fecha_inicio_texto, fecha_fin_texto)
            )
            reservas = mi_cursor.fetchall()

            if not reservas:
                print("\nNo hay reservas en este rango.\n")
                return
            else:
                lista_mostrar = [
                    [folio, evento, datetime.datetime.strptime(fecha, "%Y-%m-%d").strftime("%m-%d-%Y")] 
                    for folio, fecha, evento in reservas
                ]
                print(tabulate(lista_mostrar, headers=["Folio", "Evento", "Fecha"], tablefmt="grid"))

            while True:
                folio_a_cancelar = input("\nIngrese el folio del evento a cancelar (o escriba 'EXIT' para salir): ").strip()
                if folio_a_cancelar.upper() == "EXIT":
                    print("\nOperación cancelada.\n")
                    return

                try:
                    folio_validacion = int(folio_a_cancelar)
                except ValueError:
                    print("\nFolio inválido. Intente de nuevo.")
                    continue

                mi_cursor.execute("SELECT folio, evento, estado FROM reserva WHERE folio = ?", (folio_validacion,))
                resultado = mi_cursor.fetchone()

                if not resultado:
                    print("\nFolio no encontrado. Intente de nuevo.")
                    continue

                folio, evento, estado = resultado
                if estado == "CANCELADA":
                    print("\nEsta reservación ya está cancelada.")
                    continue

                while True:
                    confirmacion = input(f"\n¿Está seguro que desea cancelar la reservación '{evento}'? (S/N): ").strip().upper()
                    
                    if confirmacion == "S":
                        mi_cursor.execute("UPDATE reserva SET estado = 'CANCELADA' WHERE folio = ?", (folio,))
                        conn.commit()
                        print(f"\nReservación '{evento}' cancelada correctamente.\n")
                        return
                    elif confirmacion == "N":
                        print("\nCancelación abortada. Volviendo al menú.\n")
                        return
                    else:
                        print("\nRespuesta no válida. Intente con 'S' o 'N'.")

    except Error as e:
        print(e)
    except Exception:
        print(f"\nSe produjo el siguiente error: {sys.exc_info()[0]}")


def main():
    while True:
        print("\n*" + "*" * 80)
        print(f"{'MENU PRINCIPAL':^78}")
        print("*" * 80)
        print("1. Registrar la reservación de una sala.")
        print("2. Editar el nombre del evento de una reservación ya hecha.")
        print("3. Consultar las reservaciones existentes para una fecha específica.")
        print("4. Cancelar una reservación.")
        print("5. Registrar a un nuevo cliente.")
        print("6. Registrar una sala.")
        print("7. Salir")

        try:
            opcion = int(input("\nSeleccione una opción: "))
        except ValueError:
            print("\nOpción incorrecta. Intente de nuevo.\n")
            continue

        if opcion in [1, 2, 3, 4]:
            try:
                with sqlite3.connect("ReservasCoworking.db") as conn:
                    cursor = conn.cursor()
                    cursor.execute("SELECT COUNT(clave) FROM clientes")
                    total_clientes = cursor.fetchall()[0][0]

                    cursor.execute("SELECT COUNT(clave) FROM salas")
                    total_salas = cursor.fetchall()[0][0]

                    if total_clientes == 0:
                        print("\nDebe registrar al menos un cliente primero.")
                        continue

                    if total_salas == 0:
                        print("\nDebe registrar al menos una sala primero.")
                        continue

            except Error as e:
                print(f"\nError al verificar la base de datos: {e}")
                continue

        if opcion == 1:
            registrar_reserva_de_sala()
        elif opcion == 2:
            editar_nombre_de_evento()
        elif opcion == 3:
            consultar_reservas_por_fecha()
        elif opcion == 4:
            cancelar_reservas()
        elif opcion == 5:
            registrar_cliente()
        elif opcion == 6:
            registrar_sala()
        elif opcion == 7:
            while True:
                confirmar_salida_menu = input("\n¿Está seguro que desea salir del programa? (S/N): ").strip().upper()

                if confirmar_salida_menu == "":
                    print("\nNo se puede dejar vacío.")
                    continue
                elif confirmar_salida_menu == "N":
                    print("\nSe ha regresado al menú principal.\n")
                    break  
                elif confirmar_salida_menu == "S":
                    print("\nSe ha cerrado el programa con éxito.\n")
                    return 
                else:
                    print("\nNo procede. Intente escribir 'S' o 'N'.")
                    continue

        else:
            print("\nOpción incorrecta.\n")

if __name__ == "__main__":
    main()
